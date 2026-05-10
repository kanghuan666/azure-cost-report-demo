import azure.functions as func
import datetime
import logging
import os
import io
import time

from azure.core.exceptions import HttpResponseError
from azure.identity import DefaultAzureCredential
from azure.mgmt.costmanagement import CostManagementClient
from azure.storage.blob import BlobServiceClient
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout

app = func.FunctionApp()


# ==========================================
# 共通スタイル定義
# ==========================================

FONT_NAME = "Meiryo"
COLOR_PRIMARY = "1F4E78"      # ダークブルー（ヘッダー）
COLOR_ACCENT = "2E75B6"       # ミディアムブルー（強調）
COLOR_LIGHT_BG = "F5F9FC"     # ゼブラの淡色
COLOR_CARD_BG = "EAF2F8"      # サマリーカードの背景
COLOR_RED = "C00000"
COLOR_GREEN = "006100"
COLOR_RED_BG = "FFCCCC"
COLOR_GREEN_BG = "CCFFCC"
COLOR_BORDER = "CCCCCC"


def _font(size=11, bold=False, color="000000", name=FONT_NAME):
    return Font(name=name, size=size, bold=bold, color=color)


def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _border():
    side = Side(border_style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


# ==========================================
# シート整形ヘルパー
# ==========================================

def beautify_sheet(ws, money_columns=None, signed_columns=None):
    """シートを整形：ヘッダー、ゼブラ、罫線、列幅自動調整、金額フォーマット"""
    money_columns = money_columns or []
    signed_columns = signed_columns or []  # +/- を表示する列

    # ヘッダー
    for cell in ws[1]:
        cell.fill = _fill(COLOR_PRIMARY)
        cell.font = _font(size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ボディ
    border = _border()
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = _font(size=10)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = _fill(COLOR_LIGHT_BG)
            col_letter = cell.column_letter
            col_header = ws[f"{col_letter}1"].value
            if col_header in signed_columns:
                cell.number_format = '¥+#,##0;¥-#,##0;¥0'
                cell.alignment = Alignment(horizontal="right")
            elif col_header in money_columns:
                cell.number_format = '¥#,##0'
                cell.alignment = Alignment(horizontal="right")

    # 列幅
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                value = str(cell.value) if cell.value is not None else ""
                length = len(value)
                if any(ord(c) > 127 for c in value):
                    length = length * 2
                max_length = max(max_length, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    ws.row_dimensions[1].height = 28


def highlight_change_rate(ws, rate_column_name='増減率 (%)'):
    """増減率列を赤・緑でハイライト"""
    rate_col_idx = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == rate_column_name:
            rate_col_idx = col_idx
            break
    if not rate_col_idx:
        return

    for row in ws.iter_rows(min_row=2, min_col=rate_col_idx, max_col=rate_col_idx):
        cell = row[0]
        if isinstance(cell.value, (int, float)):
            if cell.value > 20:
                cell.fill = _fill(COLOR_RED_BG)
                cell.font = _font(size=10, bold=True, color=COLOR_RED)
            elif cell.value < -20:
                cell.fill = _fill(COLOR_GREEN_BG)
                cell.font = _font(size=10, bold=True, color=COLOR_GREEN)
            cell.number_format = '+0.0"%";-0.0"%";"-"'

def add_bar_chart(ws, title, data_col, label_col, chart_anchor="D2", top_n=10):
    """シートに横棒グラフを追加（サービス名・金額両方表示）"""
    max_row = min(ws.max_row, top_n + 1)
    if max_row < 2:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = title
    chart.height = 10
    chart.width = 20
    chart.legend = None  # 凡例不要（Y軸にサービス名がある）

    # 軸タイトルは消すが、軸ラベル（カテゴリ名）は残す
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    data_ref = Reference(ws, min_col=data_col, min_row=1, max_row=max_row, max_col=data_col)
    cats_ref = Reference(ws, min_col=label_col, min_row=2, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    chart.dLbls = DataLabelList(
        showVal=True,
        showCatName=False,
        showSerName=False,
        showLegendKey=False,
    )

    ws.add_chart(chart, chart_anchor)

def write_cover_sheet(wb, period_label, prev_period_label, total_current, total_previous):
    """先頭に概要シートを追加（カードレイアウト）"""
    cover = wb.create_sheet('レポート概要', 0)

    delta = total_current - total_previous
    delta_rate = (delta / total_previous * 100) if total_previous != 0 else 0

    # タイトル
    cover.merge_cells('B2:E3')
    cover['B2'] = '📊 Azure 月次コストレポート'
    cover['B2'].font = _font(size=22, bold=True, color=COLOR_PRIMARY)
    cover['B2'].alignment = Alignment(horizontal="left", vertical="center")

    # メタ情報
    meta_rows = [
        ('対象期間', period_label),
        ('生成日時', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for i, (label, value) in enumerate(meta_rows, start=5):
        cover[f'B{i}'] = label
        cover[f'C{i}'] = value
        cover[f'B{i}'].font = _font(size=11, bold=True)
        cover[f'C{i}'].font = _font(size=11)

    # サマリーカード（先月）
    cover.merge_cells('B8:C8')
    cover['B8'] = f'先月 ({period_label}) 合計コスト'
    cover['B8'].font = _font(size=10, bold=True, color="555555")
    cover['B8'].fill = _fill(COLOR_CARD_BG)
    cover['B8'].alignment = Alignment(horizontal="center", vertical="center")

    cover.merge_cells('B9:C10')
    cover['B9'] = total_current
    cover['B9'].font = _font(size=24, bold=True, color=COLOR_PRIMARY)
    cover['B9'].number_format = '¥#,##0'
    cover['B9'].alignment = Alignment(horizontal="center", vertical="center")
    cover['B9'].fill = _fill(COLOR_CARD_BG)
    cover['C10'].fill = _fill(COLOR_CARD_BG)

    # サマリーカード（先々月）
    cover.merge_cells('D8:E8')
    cover['D8'] = f'先々月 ({prev_period_label}) 合計コスト'
    cover['D8'].font = _font(size=10, bold=True, color="555555")
    cover['D8'].fill = _fill(COLOR_CARD_BG)
    cover['D8'].alignment = Alignment(horizontal="center", vertical="center")

    cover.merge_cells('D9:E10')
    cover['D9'] = total_previous
    cover['D9'].font = _font(size=24, bold=True, color="555555")
    cover['D9'].number_format = '¥#,##0'
    cover['D9'].alignment = Alignment(horizontal="center", vertical="center")
    cover['D9'].fill = _fill(COLOR_CARD_BG)
    cover['E10'].fill = _fill(COLOR_CARD_BG)

    # 差額・増減率カード
    delta_color = COLOR_RED if delta > 0 else (COLOR_GREEN if delta < 0 else "555555")

    cover.merge_cells('B12:C12')
    cover['B12'] = '差額'
    cover['B12'].font = _font(size=10, bold=True, color="555555")
    cover['B12'].fill = _fill(COLOR_CARD_BG)
    cover['B12'].alignment = Alignment(horizontal="center", vertical="center")

    cover.merge_cells('B13:C14')
    cover['B13'] = delta
    cover['B13'].font = _font(size=20, bold=True, color=delta_color)
    cover['B13'].number_format = '¥+#,##0;¥-#,##0;¥0'
    cover['B13'].alignment = Alignment(horizontal="center", vertical="center")
    cover['B13'].fill = _fill(COLOR_CARD_BG)
    cover['C14'].fill = _fill(COLOR_CARD_BG)

    cover.merge_cells('D12:E12')
    cover['D12'] = '増減率'
    cover['D12'].font = _font(size=10, bold=True, color="555555")
    cover['D12'].fill = _fill(COLOR_CARD_BG)
    cover['D12'].alignment = Alignment(horizontal="center", vertical="center")

    cover.merge_cells('D13:E14')
    cover['D13'] = delta_rate / 100
    cover['D13'].font = _font(size=20, bold=True, color=delta_color)
    cover['D13'].number_format = '+0.0%;-0.0%;0.0%'
    cover['D13'].alignment = Alignment(horizontal="center", vertical="center")
    cover['D13'].fill = _fill(COLOR_CARD_BG)
    cover['E14'].fill = _fill(COLOR_CARD_BG)

    # 警告メッセージ
    cover.merge_cells('B16:E17')
    if delta_rate > 20:
        cover['B16'] = '⚠️  警告：先月のコストが先々月比で 20% 以上増加しています。詳細をご確認ください。'
        cover['B16'].font = _font(size=12, bold=True, color=COLOR_RED)
        cover['B16'].fill = _fill(COLOR_RED_BG)
    elif delta_rate < -20:
        cover['B16'] = '✅  良好：先月のコストが先々月比で 20% 以上減少しています。'
        cover['B16'].font = _font(size=12, bold=True, color=COLOR_GREEN)
        cover['B16'].fill = _fill(COLOR_GREEN_BG)
    else:
        cover['B16'] = 'ℹ️  コストは前月と概ね同水準です（変動 20% 以内）。'
        cover['B16'].font = _font(size=12, color="555555")
        cover['B16'].fill = _fill(COLOR_LIGHT_BG)
    cover['B16'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ナビゲーション
    cover['B19'] = '📑 シート一覧'
    cover['B19'].font = _font(size=11, bold=True, color=COLOR_PRIMARY)
    nav_items = [
        ('Summary', 'サービス別コスト + Top 10 グラフ'),
        ('環比対比', '先月 vs 先々月の比較・増減率ハイライト'),
    ]
    for i, (sheet_name, description) in enumerate(nav_items, start=20):
        cover[f'B{i}'] = f'• {sheet_name}'
        cover[f'C{i}'] = description
        cover[f'B{i}'].font = _font(size=10, bold=True)
        cover[f'C{i}'].font = _font(size=10, color="555555")
        cover.merge_cells(f'C{i}:E{i}')

    # 列幅・行高
    cover.column_dimensions['A'].width = 2
    cover.column_dimensions['B'].width = 22
    cover.column_dimensions['C'].width = 22
    cover.column_dimensions['D'].width = 22
    cover.column_dimensions['E'].width = 22

    for row_num in [8, 12]:
        cover.row_dimensions[row_num].height = 22
    for row_num in [9, 13]:
        cover.row_dimensions[row_num].height = 28
    cover.row_dimensions[16].height = 30
    cover.row_dimensions[17].height = 18

    cover.sheet_view.showGridLines = False  # グリッド線非表示でカード感UP


# ==========================================
# データ取得
# ==========================================

def query_cost(cost_client, scope, start_date, end_date):
    """指定期間のコストをサービス別に取得（429リトライ付き）"""
    start_dt = datetime.datetime.combine(start_date, datetime.time.min)
    end_dt = datetime.datetime.combine(end_date, datetime.time.max)

    query = {
        "type": "ActualCost",
        "timeframe": "Custom",
        "timePeriod": {
            "from": start_dt.isoformat(),
            "to": end_dt.isoformat()
        },
        "dataset": {
            "granularity": "None",
            "aggregation": {
                "totalCost": {"name": "Cost", "function": "Sum"}
            },
            "grouping": [
                {"type": "Dimension", "name": "ServiceName"}
            ]
        }
    }

    max_retries = 5
    for attempt in range(max_retries):
        try:
            result = cost_client.query.usage(scope=scope, parameters=query)
            rows = list(result.rows) if result.rows else []
            columns = [c.name for c in result.columns]
            return pd.DataFrame(rows, columns=columns)
        except HttpResponseError as e:
            if e.status_code == 429 and attempt < max_retries - 1:
                wait_time = 30 * (attempt + 1)
                logging.warning(f'429 - {wait_time}秒待機して再試行（{attempt + 1}/{max_retries}）')
                time.sleep(wait_time)
            else:
                raise
    return pd.DataFrame()


# ==========================================
# メイン関数
# ==========================================

@app.timer_trigger(schedule="0 0 9 1 * *",
                   arg_name="myTimer",
                   run_on_startup=False,
                   use_monitor=False)
def monthlyCostReport(myTimer: func.TimerRequest) -> None:
    logging.info('月次コストレポート生成を開始します')

    subscription_id = os.environ["SUBSCRIPTION_ID"]
    storage_account = os.environ["STORAGE_ACCOUNT_NAME"]
    container_name = os.environ.get("REPORT_CONTAINER_NAME", "reports")

    # 期間計算
    today = datetime.date.today()
    first_of_this_month = today.replace(day=1)
    last_month_end = first_of_this_month - datetime.timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    period_label = last_month_start.strftime('%Y-%m')

    prev_month_end = last_month_start - datetime.timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)
    prev_period_label = prev_month_start.strftime('%Y-%m')

    logging.info(f'期間: 先月 {period_label}, 先々月 {prev_period_label}')

    # 認証 + データ取得
    credential = DefaultAzureCredential()
    cost_client = CostManagementClient(credential)
    scope = f"/subscriptions/{subscription_id}"

    df_current = query_cost(cost_client, scope, last_month_start, last_month_end)
    df_previous = query_cost(cost_client, scope, prev_month_start, prev_month_end)

    # ↓↓↓ ローカルテスト用モックデータ（本番デプロイ時はコメントアウト）
    # df_current = pd.DataFrame({
    #     'ServiceName': ['Azure Functions', 'Storage', 'Application Insights', 'Cost Management', 'Monitor'],
    #     'Cost': [1230, 540, 320, 0, 90]
    # })
    # df_previous = pd.DataFrame({
    #     'ServiceName': ['Azure Functions', 'Storage', 'Application Insights'],
    #     'Cost': [800, 600, 200]
    # })
    # ↑↑↑ ここまで

    logging.info(f'先月: {len(df_current)}行, 先々月: {len(df_previous)}行')

    # ¥0 のサービスは除外（ノイズになるため）
    if not df_current.empty:
        df_current = df_current[df_current['Cost'] > 0].reset_index(drop=True)
    if not df_previous.empty:
        df_previous = df_previous[df_previous['Cost'] > 0].reset_index(drop=True)

    # Excel生成
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if df_current.empty and df_previous.empty:
            pd.DataFrame({"メッセージ": ["対象期間にコストデータがありません"]}).to_excel(
                writer, sheet_name='Summary', index=False)
        else:
            current_summary = (df_current.groupby('ServiceName')['Cost'].sum().reset_index()
                               if not df_current.empty
                               else pd.DataFrame(columns=['ServiceName', 'Cost']))
            previous_summary = (df_previous.groupby('ServiceName')['Cost'].sum().reset_index()
                                if not df_previous.empty
                                else pd.DataFrame(columns=['ServiceName', 'Cost']))

            # 環比対比
            comparison = current_summary.merge(
                previous_summary, on='ServiceName', how='outer', suffixes=('_curr', '_prev')
            ).fillna(0)
            comparison.columns = ['サービス名', f'{period_label} (¥)', f'{prev_period_label} (¥)']
            comparison['差額'] = comparison[f'{period_label} (¥)'] - comparison[f'{prev_period_label} (¥)']
            comparison['増減率 (%)'] = comparison.apply(
                lambda r: ((r[f'{period_label} (¥)'] - r[f'{prev_period_label} (¥)']) / r[f'{prev_period_label} (¥)'] * 100)
                if r[f'{prev_period_label} (¥)'] != 0 else 0,
                axis=1
            ).round(1)
            comparison = comparison.sort_values(f'{period_label} (¥)', ascending=False)
            comparison.to_excel(writer, sheet_name='環比対比', index=False)

            # サービス別サマリー（先月のみ、降順）
            if not current_summary.empty:
                summary_jp = current_summary.copy()
                summary_jp.columns = ['サービス名', 'コスト (¥)']
                summary_jp = summary_jp.sort_values('コスト (¥)', ascending=False).reset_index(drop=True)
                summary_jp.to_excel(writer, sheet_name='Summary', index=False)

    output.seek(0)
    wb = load_workbook(output)

    # 美化
    if 'Summary' in wb.sheetnames:
        beautify_sheet(wb['Summary'], money_columns=['コスト (¥)'])
        add_bar_chart(wb['Summary'],
                      title=f"サービス別コスト Top 10 ({period_label})",
                      data_col=2, label_col=1)

    if '環比対比' in wb.sheetnames:
        beautify_sheet(
            wb['環比対比'],
            money_columns=[f'{period_label} (¥)', f'{prev_period_label} (¥)'],
            signed_columns=['差額'],
        )
        highlight_change_rate(wb['環比対比'])

    # 封面シート
    if not df_current.empty or not df_previous.empty:
        total_current = float(df_current['Cost'].sum()) if not df_current.empty else 0
        total_previous = float(df_previous['Cost'].sum()) if not df_previous.empty else 0
        write_cover_sheet(wb, period_label, prev_period_label, total_current, total_previous)

    # 再保存
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # アップロード
    blob_url = f"https://{storage_account}.blob.core.windows.net"
    blob_service = BlobServiceClient(account_url=blob_url, credential=credential)
    container_client = blob_service.get_container_client(container_name)
    try:
        container_client.create_container()
        logging.info(f'コンテナを作成: {container_name}')
    except Exception:
        pass

    blob_name = f"cost-report-{period_label}.xlsx"
    blob_client = blob_service.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(final_output.getvalue(), overwrite=True)

    logging.info(f'レポート出力完了: {container_name}/{blob_name}')